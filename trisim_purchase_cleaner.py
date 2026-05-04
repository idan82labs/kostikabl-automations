# -*- coding: utf-8 -*-
"""
trisim_purchase_cleaner_v1_4h44_FIX_HEBREW_QUOTES.py

FIXES IN THIS VERSION:
- v1.4h44: CRITICAL FIX - Hebrew abbreviation quotes (מ"מ, מ"ר, ק"ג etc.) were being
  treated as CSV quote delimiters by _sanitize_illegal_newlines, causing it to merge
  dozens of rows into one mega-line (714+ fields). Now uses _has_quotes_in_data to
  detect real CSV quoting vs Hebrew abbreviations, and uses csv.QUOTE_NONE when
  the data only contains Hebrew-style quotes.
- v1.4h43: Rows with doubled fields are SPLIT into multiple rows instead of truncated.
"""

import sys, io, csv, re, tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
import numpy as np, pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter

APP_VERSION = "1.4h47-PICKER_LAYOUT_FIX"

# Create a log file for debugging
def log_message(msg):
    try:
        log_path = Path(sys.executable if getattr(sys, "frozen", False) else __file__).parent / "debug_log.txt"
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.now()}] {msg}\n")
    except:
        pass

# ---------- ENCODING & CSV/XML CONVERSION ----------

ENCODINGS_PRIMARY = ["utf-8-sig","utf-8"]
ENCODINGS_HEBREW  = ["cp1255","windows-1255","iso-8859-8"]
ENCODINGS_FALLBK  = ["windows-1252","latin-1"]

def _decode_best(raw: bytes) -> str:
    # Try UTF first, then Hebrew codepages, then safe fallbacks.
    for bank in (ENCODINGS_PRIMARY, ENCODINGS_HEBREW, ENCODINGS_FALLBK):
        for enc in bank:
            try:
                txt = raw.decode(enc)
                # Heuristic: if we see many mojibake chars, try next bank
                bad = txt.count("�") + txt.count("™") + txt.count("�")
                if bank is ENCODINGS_PRIMARY and bad > 10:
                    # looks wrong for UTF – try Hebrew explicitly
                    continue
                log_message(f"Successfully decoded with {enc}")
                return txt
            except Exception as e:
                log_message(f"Failed to decode with {enc}: {e}")
                pass
    # Last resort: decode ignoring errors
    log_message("Using fallback UTF-8 with ignore")
    return raw.decode("utf-8", "ignore")

def _looks_xml(s: str) -> bool:
    snip = s.lstrip()[:200].lower()
    return snip.startswith("<") or "<?xml" in snip or "<row" in snip or "<table" in snip

def _strip_bom_and_normalize_newlines(s: str) -> str:
    return s.replace("\r\n","\n").replace("\r","\n").lstrip("\ufeff")

def _best_delimiter(sample: str) -> str:
    # Prefer semicolon like Excel UI, but compute counts ignoring quotes.
    first_k = "\n".join(sample.split("\n")[:200])
    def count_unquoted(text, ch):
        cnt=0; q=False
        for c in text:
            if c=='"': q = not q
            elif c==ch and not q: cnt += 1
        return cnt
    c_sc = count_unquoted(first_k, ';')
    c_cm = count_unquoted(first_k, ',')
    c_tb = count_unquoted(first_k, '\t')
    # If clear winner -> choose; else bias to semicolon.
    winners = sorted([(c_sc,';'),(c_cm,','),(c_tb,'\t')], reverse=True)
    if winners[0][0] >= max(2, winners[1][0]*1.2):  # clear lead
        log_message(f"Delimiter detected: '{winners[0][1]}' (count: {winners[0][0]})")
        return winners[0][1]
    log_message("Delimiter: using semicolon bias")
    return ';'  # bias

def _fix_leading_trailing_delims(text: str, sep: str) -> str:
    # Remove a systematic leading delimiter and trim trailing extra separators.
    lines = text.split("\n")
    nz = [ln for ln in lines if ln.strip()]
    if not nz: return text
    lead_ratio = sum(1 for ln in nz if ln.startswith(sep)) / max(1,len(nz))
    if lead_ratio >= 0.6:
        lines = [ln[1:] if ln.startswith(sep) else ln for ln in lines]
    # Strip repeated trailing separators - ENHANCED to handle single trailing sep too
    if sep in (';',',','\t','|'):
        lines = [re.sub(rf"{re.escape(sep)}+$","",ln) for ln in lines]

    # Recompute nz after stripping
    nz = [ln for ln in lines if ln.strip()]

    # NEW: Additional check for field count mismatch between header and data
    # If data rows consistently have more fields than header, strip last field from data ONLY
    if len(nz) > 1:
        header_fields = nz[0].count(sep) + 1
        data_field_counts = [ln.count(sep) + 1 for ln in nz[1:20]]  # Sample first 20 data rows
        if data_field_counts and min(data_field_counts) > header_fields:
            # Data rows have more fields - likely trailing empty field
            log_message(f"Field mismatch detected: header has {header_fields} fields, data has {min(data_field_counts)}-{max(data_field_counts)} fields")
            # FIXED: Only strip trailing separators from DATA rows, not header
            # Keep header as-is, strip from all other lines
            fixed_lines = [lines[0]]  # Header unchanged
            for ln in lines[1:]:
                fixed_lines.append(ln.rstrip(sep) if ln.strip() else ln)
            lines = fixed_lines
            log_message(f"Stripped trailing separators from data rows only (not header)")

    return "\n".join(lines)

def _sanitize_illegal_newlines(text: str) -> str:
    # Replace unquoted newlines inside fields by spaces (rare export glitch)
    out=[]; q=False
    for ch in text:
        if ch == '"':
            q = not q
            out.append(ch)
        elif ch == '\n' and q:
            out.append(' ')  # keep row intact
        else:
            out.append(ch)
    return "".join(out)

def _has_quotes_in_data(text: str, delim: str = ';') -> bool:
    """Check if the CSV data contains REAL quoted fields (not just quotes in text like פ"ת)

    Real CSV quoting looks like: ;"field value"; or ,"field value",
    NOT like: שדרות הסביונים פ"ת (quotes embedded in Hebrew text)
    """
    if '"' not in text:
        return False

    # Check for real CSV quoting patterns:
    # 1. Delimiter followed by quote: ;" or ,"
    # 2. Quote followed by delimiter: "; or ",
    # 3. Line starting with quote (first field quoted)
    import re

    # Pattern for real CSV quoted fields
    real_quote_patterns = [
        rf'{re.escape(delim)}"',  # delimiter followed by quote: ;"
        rf'"{re.escape(delim)}',  # quote followed by delimiter: ";
        r'^"',                     # line starts with quote
        r'\n"',                    # quote at start of line
    ]

    for pattern in real_quote_patterns:
        if re.search(pattern, text, re.MULTILINE):
            return True

    return False

def read_any_to_dataframe_and_write_utf8(src: Path):
    log_message(f"=== Reading file: {src} ===")
    raw = src.read_bytes()
    log_message(f"Raw bytes: {len(raw)}")

    decoded = _decode_best(raw)
    decoded = _strip_bom_and_normalize_newlines(decoded)

    if _looks_xml(decoded):
        log_message("File format: XML")
        # XML path
        xml_text = _sanitize_illegal_newlines(decoded)
        df = pd.read_xml(io.StringIO(xml_text))
        # Write UTF8 copy
        utf_p = src.with_name(f"{src.stem} (utf8).csv")
        df.to_csv(utf_p, index=False, encoding="utf-8")
        log_message(f"Saved UTF-8 copy: {utf_p}")
        return df, utf_p

    # CSV path
    log_message("File format: CSV")
    delim = _best_delimiter(decoded)

    # v1.4h44: Only sanitize newlines inside quotes if the data has REAL CSV quoting.
    # Hebrew abbreviations like מ"מ (mm), מ"ר (m²), ק"ג (kg) contain literal quotes
    # that are NOT CSV quoting.  _sanitize_illegal_newlines treats them as open-quotes,
    # joining dozens of rows into one mega-line and destroying the data.
    has_real_quoting = _has_quotes_in_data(decoded, delim)
    log_message(f"Has real CSV quoting: {has_real_quoting}")
    if has_real_quoting:
        cleaned = _fix_leading_trailing_delims(_sanitize_illegal_newlines(decoded), delim)
    else:
        cleaned = _fix_leading_trailing_delims(decoded, delim)

    log_message(f"Using simple delimiter: {delim} (regex disabled for reliability)")
    engine = "c"
    sep = delim

    # ULTIMATE FIX FOR COLUMN SHIFT:
    # pandas skips rows with mismatched field counts, losing data entirely!
    # Solution: Manually parse and fix field counts BEFORE pandas sees it

    lines = cleaned.split('\n')
    log_message(f"Total lines after split: {len(lines)}")
    if len(lines) < 2:
        raise RuntimeError("קובץ ריק או פגום")

    header = lines[0]
    expected_fields = header.count(sep) + 1
    log_message(f"Header has {expected_fields} fields")

    # Check if data rows have different field counts
    data_lines = [ln for ln in lines[1:] if ln.strip()]
    log_message(f"Non-empty data lines: {len(data_lines)}")

    if data_lines:
        data_field_counts = [ln.count(sep) + 1 for ln in data_lines[:20]]
        max_fields = max(data_field_counts) if data_field_counts else expected_fields
        min_fields = min(data_field_counts) if data_field_counts else expected_fields

        log_message(f"Sample field counts (first 20 rows): min={min_fields}, max={max_fields}")

        if max_fields > expected_fields:
            log_message(f"⚠️ Column shift detected: data has up to {max_fields} fields, header has {expected_fields}")

            # NEW v1.4h43: Check if fields are approximately DOUBLED (concatenated rows)
            # If field count is ~2x expected, SPLIT rows instead of truncating
            if max_fields >= expected_fields * 1.8:  # At least 1.8x (e.g., 34 vs 17)
                log_message(f"🔧 DETECTED CONCATENATED ROWS: {max_fields} fields ≈ 2 × {expected_fields}")
                log_message("Splitting concatenated rows into separate rows...")

                fixed_lines = [header]
                split_count = 0

                for ln in data_lines:
                    fields = ln.split(sep)
                    field_count = len(fields)

                    if field_count >= expected_fields * 1.8:
                        # This line has doubled fields - split it
                        # First logical row: fields 0 to expected_fields-1
                        row1 = sep.join(fields[:expected_fields])
                        fixed_lines.append(row1)

                        # Second logical row: fields expected_fields to 2*expected_fields-1
                        remaining = fields[expected_fields:]
                        if len(remaining) >= expected_fields:
                            row2 = sep.join(remaining[:expected_fields])
                            fixed_lines.append(row2)
                            split_count += 1

                            # Check for third row (rare but possible)
                            remaining2 = remaining[expected_fields:]
                            if len(remaining2) >= expected_fields:
                                row3 = sep.join(remaining2[:expected_fields])
                                fixed_lines.append(row3)
                                log_message(f"Found 3+ concatenated rows in one line!")
                        elif len(remaining) > 0:
                            # Partial second row - pad with empty fields
                            while len(remaining) < expected_fields:
                                remaining.append('')
                            row2 = sep.join(remaining[:expected_fields])
                            fixed_lines.append(row2)
                            split_count += 1
                    elif field_count > expected_fields:
                        # Slightly more than expected but not doubled - truncate
                        fixed_ln = sep.join(fields[:expected_fields])
                        fixed_lines.append(fixed_ln)
                    elif field_count < expected_fields:
                        # Too few fields - pad
                        while len(fields) < expected_fields:
                            fields.append('')
                        fixed_lines.append(sep.join(fields))
                    else:
                        # Perfect field count
                        fixed_lines.append(ln)

                cleaned = '\n'.join(fixed_lines)
                log_message(f"✓ Split {split_count} concatenated rows")
                log_message(f"✓ Now have {len(fixed_lines) - 1} data rows (was {len(data_lines)})")
            else:
                # Not doubled, just extra trailing fields - truncate as before
                log_message("Fixing by truncating extra fields...")

                fixed_lines = [header]
                truncate_count = 0
                for ln in data_lines:
                    fields = ln.split(sep)
                    if len(fields) > expected_fields:
                        fixed_ln = sep.join(fields[:expected_fields])
                        fixed_lines.append(fixed_ln)
                        truncate_count += 1
                    else:
                        fixed_lines.append(ln)

                cleaned = '\n'.join(fixed_lines)
                log_message(f"✓ Truncated {truncate_count} rows with extra fields")

    # Now read with pandas
    # v1.4h44: disable quoting when data only has Hebrew abbreviation quotes (מ"מ etc.)
    quote_kwargs = dict(quotechar='"') if has_real_quoting else dict(quoting=csv.QUOTE_NONE)
    try:
        df = pd.read_csv(
            io.StringIO(cleaned),
            engine=engine,
            sep=sep,
            **quote_kwargs,
            skipinitialspace=True,
            keep_default_na=False,
            dtype=str
        )
    except Exception as e:
        log_message(f"CSV parsing failed: {e}")
        log_message("Retrying with robust manual parsing...")

        # IMPROVED v39: Manually fix field mismatches before pandas sees them
        lines = cleaned.split('\n')
        if len(lines) < 2:
            raise RuntimeError("קובץ ריק או פגום")

        header = lines[0].strip()
        expected_fields = header.count(delim) + 1
        log_message(f"Expected {expected_fields} fields per row")

        fixed_lines = [header]
        skipped_count = 0
        fixed_count = 0

        for i, line in enumerate(lines[1:], start=2):
            if not line.strip():
                continue

            field_count = line.count(delim) + 1

            if field_count == expected_fields:
                # Perfect - keep as is
                fixed_lines.append(line)
            elif field_count > expected_fields:
                # Too many fields - truncate to expected count
                fields = line.split(delim)
                fixed_line = delim.join(fields[:expected_fields])
                fixed_lines.append(fixed_line)
                fixed_count += 1
                if fixed_count <= 5:  # Log first 5 only
                    log_message(f"⚠️ Line {i}: Fixed {field_count} → {expected_fields} fields")
            else:
                # Too few fields - pad with empty strings
                fields = line.split(delim)
                while len(fields) < expected_fields:
                    fields.append('')
                fixed_line = delim.join(fields)
                fixed_lines.append(fixed_line)
                fixed_count += 1
                if fixed_count <= 5:  # Log first 5 only
                    log_message(f"⚠️ Line {i}: Padded {field_count} → {expected_fields} fields")

        if fixed_count > 0:
            log_message(f"✓ Fixed {fixed_count} rows with field mismatches")

        cleaned_fixed = '\n'.join(fixed_lines)

        # Now try parsing the fixed data
        try:
            df = pd.read_csv(
                io.StringIO(cleaned_fixed),
                engine="python",
                sep=delim,
                **quote_kwargs,
                skipinitialspace=True,
                keep_default_na=False,
                dtype=str
            )
            log_message(f"✓ Manual parsing succeeded!")
        except Exception as e2:
            log_message(f"✗ Manual parsing also failed: {e2}")
            log_message("Falling back to skip mode as last resort...")
            df = pd.read_csv(
                io.StringIO(cleaned),
                engine="python",
                sep=delim,
                **quote_kwargs,
                skipinitialspace=True,
                keep_default_na=False,
                on_bad_lines='skip',
                encoding_errors='ignore',
                dtype=str
            )

    # Column cleanup
    df.columns = [str(c).replace('\ufeff','').replace('\u00A0',' ').strip() for c in df.columns]
    log_message(f"Loaded {len(df)} rows, {len(df.columns)} columns")
    log_message(f"Columns: {list(df.columns)}")

    # Save a clean UTF-8 copy
    utf_p = src.with_name(f"{src.stem} (utf8).csv")
    utf_p.write_text(cleaned, encoding="utf-8", newline="\n")
    log_message(f"Saved UTF-8 copy: {utf_p}")
    return df, utf_p

# ---------- HEADERS NORMALIZATION, MAPPING, UI, CALC ----------
def norm_header(s: str) -> str:
    s = str(s).replace('\ufeff','').replace('\u00A0',' ')
    s = s.replace('"',' ').replace('״',' ').replace("'", " ")
    s = re.sub(r'[^A-Za-z0-9\u0590-\u05FF ]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def normalize_csv_columns(df: pd.DataFrame) -> pd.DataFrame:
    cols = list(df.columns)
    nm = {norm_header(c): c for c in cols}
    def pick(*aliases):
        for a in aliases:
            if norm_header(a) in nm:
                return nm[norm_header(a)]
        return None
    rename = {}
    m = {
      "project":   pick("פרויקט","project"),
      "sku_desc":  pick("שם-מקייט","שם מק ט","שם מקט","שם מק\"ט"),
      "desc":      pick("תאור","תיאור","תיאור פריט"),
      "width":     pick("רוחב","רוחב במ מ","רוחב בממ","width"),
      "height":    pick("גובה","גובה במ מ","גובה בממ","height"),
      "area":      pick("מידה","שטח ליחידה במ ר","שטח ליחידה במ\"ר","area"),
      "code":      pick("קוד זיהוי","מס פרט","מספר פרט","מספר פריט","קוד פריט","קוד פנימי"),
      "vendor":    pick("ספק","יצרן","vendor"),
      "color":     pick("צבע חיצוני","צבע פנימי","גוון אלומיניום","גוון"),
    }
    mapping = {
      m["project"]: "פרויקט",
      m["sku_desc"]: "שם-מקייט",
      m["desc"]: "תאור",
      m["width"]: "רוחב",
      m["height"]: "גובה",
      m["area"]: "מידה",
      m["code"]: "קוד זיהוי",
      m["vendor"]: "ספק",
      m["color"]: "צבע חיצוני",
    }
    for src, dst in mapping.items():
        if src and src != dst:
            rename[src] = dst
    if rename:
        df = df.rename(columns=rename)
        log_message(f"Renamed columns: {rename}")
    return df

def project_from_stem(stem: str) -> str:
    s = re.sub(r"[^A-Za-z\u0590-\u05FF ]+", " ", stem)
    s = re.sub(r"\d+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s or stem

def norm_sku(x) -> str:
    return re.sub(r'[^0-9A-Za-z]+','', str(x)).lower()

def ask_for_csv() -> Path:
    root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
    p = filedialog.askopenfilename(
        title="בחר קובץ CSV/XML",
        filetypes=[("All", "*.*"), ("CSV", "*.csv"), ("XML", "*.xml")],
        parent=root
    )
    root.destroy()
    return Path(p) if p else None

def ask_for_output_dir(default_dir: Path = None) -> Path:
    """Ask user to select output directory for the cleaned file"""
    root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
    initial_dir = str(default_dir) if default_dir else None
    p = filedialog.askdirectory(
        title="בחר תיקייה לשמירת הפלט (או לחץ Cancel לשמירה במיקום המקורי)",
        initialdir=initial_dir,
        parent=root
    )
    root.destroy()
    return Path(p) if p else None

def find_mapping(exe_dir: Path, csv_parent: Path):
    possibilities = [
        exe_dir / "data" / "מקטים תריסים כולל.xlsx",
        exe_dir / "מקטים תריסים כולל.xlsx",
        csv_parent / "מקטים תריסים כולל.xlsx",
    ]
    for p in possibilities:
        if p.exists():
            log_message(f"Found mapping file: {p}")
            return p
    raise FileNotFoundError(f"לא נמצא קובץ המיפוי 'מקטים תריסים כולל.xlsx' באף אחד מהמיקומים:\n" +
                            "\n".join(str(pp) for pp in possibilities))

def read_mapping_autodetect(path: Path) -> pd.DataFrame:
    log_message(f"Reading mapping file: {path}")
    df = pd.read_excel(path, engine="openpyxl")
    df.columns = [str(c).replace('\ufeff','').replace('\u00A0',' ').strip() for c in df.columns]
    log_message(f"Mapping loaded: {len(df)} rows, columns: {list(df.columns)}")
    return df

def pick_mapping_columns(mp: pd.DataFrame):
    candidates_sku  = ["שם-מקייט","מק\"ט","מקט","sku","code"]
    candidates_desc = ["תיאור","תאור","description","desc"]
    candidates_vend = ["ספק","יצרן","vendor","supplier"]
    candidates_split= ["חלוקה","שלב משוך תחתון","bottom"]
    def find_col(cands, fallback=None):
        for c in mp.columns:
            for cand in cands:
                if cand in c or norm_header(cand) == norm_header(c):
                    return c
        return fallback
    sku_col   = find_col(candidates_sku, "שם-מקייט")
    desc_col  = find_col(candidates_desc, "תאור")
    vendor_col= find_col(candidates_vend, None)
    split_col = find_col(candidates_split, None)
    log_message(f"Mapping columns - SKU: {sku_col}, DESC: {desc_col}, VENDOR: {vendor_col}")
    return sku_col, desc_col, vendor_col, split_col

def select_codes_from_list(codes_list):
    root = tk.Tk(); root.title("בחר קודי זיהוי לשלב משוך תחתון")
    root.attributes("-topmost", True)
    w, h = 500, 600
    root.geometry(f"{w}x{h}")
    chosen = []
    def on_ok():
        chosen.clear()
        chosen.extend([c for c,v in checks.items() if v.get()])
        root.destroy()
    checks = {}
    canvas = tk.Canvas(root, bg="#f0f0f0")
    scrollbar = tk.Scrollbar(root, orient="vertical", command=canvas.yview)
    scrollable_frame = tk.Frame(canvas, bg="#f0f0f0")
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    tk.Label(scrollable_frame, text="בחר קודי זיהוי לחישוב שלב משוך תחתון:",
             font=("Arial", 11, "bold"), bg="#f0f0f0").pack(pady=10)
    for code in codes_list:
        v = tk.BooleanVar()
        checks[code] = v
        tk.Checkbutton(scrollable_frame, text=code, variable=v, font=("Arial", 10),
                       bg="#f0f0f0", anchor="w").pack(fill="x", padx=10)
    def select_all():
        for v in checks.values(): v.set(True)
    def clear_all():
        for v in checks.values(): v.set(False)

    # Reserve the bottom strip for buttons FIRST so the canvas can't eat the space.
    btn_row = tk.Frame(root, bg="#f0f0f0")
    btn_row.pack(side="bottom", fill="x", pady=10)
    tk.Button(btn_row, text="אישור", command=on_ok, font=("Arial", 11, "bold"),
              bg="#4CAF50", fg="white", padx=20, pady=5).pack(side="right", padx=6)
    tk.Button(btn_row, text="נקה הכל", command=clear_all, font=("Arial", 10),
              bg="#9E9E9E", fg="white", padx=12, pady=4).pack(side="right", padx=4)
    tk.Button(btn_row, text="בחר הכל", command=select_all, font=("Arial", 10),
              bg="#2196F3", fg="white", padx=12, pady=4).pack(side="right", padx=4)

    canvas.pack(side="left", fill="both", expand=True, padx=5, pady=5)
    scrollbar.pack(side="right", fill="y")
    root.mainloop()
    log_message(f"Selected codes for bottom rail: {chosen}")
    return chosen

def to_mm_numeric(s: pd.Series) -> pd.Series:
    def parse_val(v):
        if pd.isna(v): return np.nan
        v = str(v).strip().replace(",","").lower().replace("mm","").replace("m","")
        try:
            f = float(v)
            return f if 1 <= f <= 1e7 else np.nan
        except:
            return np.nan
    return s.map(parse_val)

def fix_units_mm(s: pd.Series) -> pd.Series:
    factor100 = (s >= 1) & (s < 100)
    s = s.copy()
    s[factor100] = (s[factor100] * 1000).round(0)
    factor10  = (s >= 100) & (s < 300)
    s[factor10]  = (s[factor10] * 10).round(0)
    factor100 = (s >= 60000) & (s <= 600000)
    s[factor100] = (s[factor100] / 100).round(0)
    return s

# ---------------- MAIN PIPELINE ----------------
def main():
    log_message("="*60)
    log_message(f"TRISIM PURCHASE CLEANER v{APP_VERSION} STARTED")
    log_message("="*60)

    src_path = ask_for_csv()
    if not src_path:
        log_message("User cancelled file selection")
        return

    # Ask for output directory (optional - defaults to same directory as input)
    output_dir = ask_for_output_dir(default_dir=src_path.parent)
    if not output_dir:
        # User cancelled or closed dialog - use same directory as input file
        output_dir = src_path.parent
        log_message(f"No output directory selected, using input file directory: {output_dir}")
    else:
        log_message(f"Output directory selected: {output_dir}")

    exe_dir = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent

    try:
        mp_path = find_mapping(exe_dir, src_path.parent)

        # NEW: universal reader that converts to UTF-8 and returns DataFrame
        df_raw, utf8_copy = read_any_to_dataframe_and_write_utf8(src_path)
        df = normalize_csv_columns(df_raw)

        # Read mapping
        mp = read_mapping_autodetect(mp_path)
        sku_col, desc_col, vendor_col, split_col = pick_mapping_columns(mp)
        mp = mp.rename(columns={sku_col:"__SKU__", desc_col:"__DESC__", (vendor_col if vendor_col else ""):"__VENDOR__"})
        if split_col and split_col in mp.columns:
            mp = mp.rename(columns={split_col: "__SPLIT__"})
        else:
            mp["__SPLIT__"] = pd.NA
        mp["__DESC_N__"] = mp["__DESC__"].astype(str).str.replace(r"\s+"," ", regex=True).str.strip()
        mp["__SKU_N__"]  = mp["__SKU__"].map(norm_sku)

        # Determine description source
        desc_source = "תאור" if ("תאור" in df.columns and df["תאור"].notna().any()) \
                      else ("שם-מקייט" if ("שם-מקייט" in df.columns and df["שם-מקייט"].notna().any()) else None)
        if not desc_source:
            raise RuntimeError(f"לא נמצאה עמודת תיאור/שם-מקייט לשיוך. כותרות: {', '.join(df.columns.astype(str))}")

        log_message(f"Using description source: {desc_source}")
        df["__DESC_N__"] = df[desc_source].astype(str).str.replace(r"\s+"," ", regex=True).str.strip()

        # NEW: Detect project-wide color BEFORE matching
        # This handles cases where only components have colors but shutters don't
        project_color = None
        for color_col in ["צבע חיצוני", "צבע פנימי", "גוון אלומיניום", "גוון"]:
            if color_col in df.columns:
                temp_series = df[color_col]
                has_data = (temp_series.notna() & (temp_series.astype(str).str.strip() != '')).any()
                if has_data:
                    # Get the most common color from the entire input file
                    non_empty = temp_series[temp_series.astype(str).str.strip() != '']
                    if len(non_empty) > 0:
                        project_color = non_empty.mode()[0] if len(non_empty.mode()) > 0 else non_empty.iloc[0]
                        log_message(f"Detected project-wide color from '{color_col}': '{project_color}' (appears in {len(non_empty)} rows)")
                        break

        if project_color:
            log_message(f"Will use project color '{project_color}' for rows without color data")
        else:
            log_message("No project-wide color detected in input file")

        # Primary join by description
        log_message(f"=== DESCRIPTION MATCHING DEBUG ===")
        log_message(f"df has {len(df)} rows")

        # Log shutter rows in CSV BEFORE matching
        shutter_mask = df["שם-מקייט"].map(norm_sku).str.startswith("130")
        shutter_rows = df[shutter_mask]
        log_message(f"Shutter rows in CSV (SKU starts with 130): {len(shutter_rows)}")
        for idx, row in shutter_rows.iterrows():
            log_message(f"  CSV idx={idx}: SKU={row.get('שם-מקייט','')}, code={row.get('קוד זיהוי','')}, desc='{row.get('__DESC_N__','')[:50]}'")

        # Log mapping descriptions for shutters
        log_message(f"Mapping shutter descriptions:")
        mp_shutters = mp[mp["__SKU_N__"].str.startswith("130")]
        for idx, row in mp_shutters.iterrows():
            log_message(f"  MAP: SKU={row['__SKU__']}, desc='{row['__DESC_N__']}'")

        j_desc = df.merge(
            mp[["__DESC_N__","__SKU__","__DESC__","__VENDOR__","__SKU_N__","__SPLIT__"]].drop_duplicates("__DESC_N__"),
            how="left", on="__DESC_N__"
        )
        matched_desc = j_desc[j_desc["__SKU__"].notna()].copy()
        log_message(f"Matched by description: {len(matched_desc)} rows")

        # Log which codes matched by description
        if 'קוד זיהוי' in matched_desc.columns:
            for idx, row in matched_desc.iterrows():
                log_message(f"  DESC MATCH idx={idx}: code={row.get('קוד זיהוי','')}, matched_SKU={row.get('__SKU__','')}")

        # FIXED: Always try SKU join too (not just if matched.empty)
        # This catches rows like מק"ט 13290003 where description differs but SKU matches
        matched_sku = pd.DataFrame()
        if "שם-מקייט" in df.columns:
            log_message("=== SKU MATCHING DEBUG ===")
            df["__SKU_N__"] = df["שם-מקייט"].map(norm_sku)
            j_sku = df.merge(
                mp[["__SKU_N__","__SKU__","__DESC__","__VENDOR__","__SPLIT__"]].drop_duplicates("__SKU_N__"),
                how="left", on="__SKU_N__"
            )
            matched_sku = j_sku[j_sku["__SKU__"].notna()].copy()
            log_message(f"Matched by SKU: {len(matched_sku)} rows")

            # Log which codes matched by SKU
            if 'קוד זיהוי' in matched_sku.columns:
                for idx, row in matched_sku.iterrows():
                    log_message(f"  SKU MATCH idx={idx}: code={row.get('קוד זיהוי','')}, matched_SKU={row.get('__SKU__','')}")

        # Combine description and SKU matches, removing duplicates
        # FIXED v1.4h38: Add detailed logging to diagnose duplicate issue
        if not matched_desc.empty and not matched_sku.empty:
            log_message(f"Before combine: matched_desc={len(matched_desc)} rows, matched_sku={len(matched_sku)} rows")
            log_message(f"matched_desc columns: {matched_desc.columns.tolist()}")
            log_message(f"matched_sku columns: {matched_sku.columns.tolist()}")
            log_message(f"df.columns (for dedup): {df.columns.tolist()}")

            # Log a sample of codes before combining
            if 'קוד זיהוי' in matched_desc.columns:
                codes_desc = matched_desc['קוד זיהוי'].dropna().unique()
                log_message(f"Codes in matched_desc: {sorted([str(c) for c in codes_desc])}")
            if 'קוד זיהוי' in matched_sku.columns:
                codes_sku = matched_sku['קוד זיהוי'].dropna().unique()
                log_message(f"Codes in matched_sku: {sorted([str(c) for c in codes_sku])}")

            # Keep original index to identify true duplicates
            matched = pd.concat([matched_desc, matched_sku], ignore_index=False)
            log_message(f"After concat: {len(matched)} rows")

            # Remove only rows that came from the same original CSV row (same index)
            matched = matched[~matched.index.duplicated(keep='first')].reset_index(drop=True)
            log_message(f"After dedup by index: {len(matched)} rows")

            # Log codes after dedup
            if 'קוד זיהוי' in matched.columns:
                codes_final = matched['קוד זיהוי'].dropna().unique()
                log_message(f"Codes after dedup: {sorted([str(c) for c in codes_final])}")
        elif not matched_desc.empty:
            matched = matched_desc
        elif not matched_sku.empty:
            matched = matched_sku
        else:
            matched = pd.DataFrame()

        # REMOVED: Auto-detection feature - only process items that are in the mapping file
        # Users should add items to the mapping file if they want them processed

        if matched.empty:
            raise RuntimeError("לא נמצאו תריסים להתאמה (לא לפי תיאור ולא לפי מק\"ט).")

        # Process dimensions
        width_mm  = fix_units_mm(to_mm_numeric(matched.get("רוחב")))
        height_mm = fix_units_mm(to_mm_numeric(matched.get("גובה")))

        # FIXED: Try to read area from "כמות" first if it's in m², otherwise from "מידה"
        # This handles files where area is in the "כמות" column (when unit is m²)
        area_m2_in = None
        if "יחידת מידה" in matched.columns:
            unit_series = matched.get("יחידת מידה").astype(str).str.lower().str.strip()
            is_area_unit = unit_series.str.contains(r'm2|מ\"ר|מר|m²', na=False, regex=True)
            if is_area_unit.any():
                # Area is in "כמות" column for rows with area units
                qty_series = pd.to_numeric(matched.get("כמות"), errors="coerce")
                area_from_qty = pd.Series([pd.NA] * len(matched), index=matched.index)
                area_from_qty[is_area_unit] = qty_series[is_area_unit]
                log_message(f"Found {is_area_unit.sum()} rows with area in 'כמות' column (unit: m²)")
                area_m2_in = area_from_qty

        # Fallback to "מידה" column if not found in "כמות"
        if area_m2_in is None or area_m2_in.isna().all():
            area_m2_in = pd.to_numeric(matched.get("מידה"), errors="coerce")
            log_message(f"Reading area from 'מידה' column")

        width_m = width_mm.astype(float)/1000.0
        height_m = height_mm.astype(float)/1000.0

        need = height_mm.isna() & area_m2_in.notna() & (width_m > 0)
        height_m = height_m.astype(object)
        # FIXED: Only calculate if there are rows that need height calculation
        # This prevents dtype error when trying to round an empty Series
        if need.any():
            height_m[need] = (area_m2_in[need] / width_m[need]).round(5)
        height_mm_final = fix_units_mm(pd.to_numeric(pd.Series(height_m).astype(float)*1000, errors="coerce"))

        # Pavel's rule: round shutter height UP to the nearest multiple of חלוקה (slat profile mm).
        # Applies only to rows where the mapping defines a positive __SPLIT__ value (50/52/55/61 mm etc.).
        if "__SPLIT__" in matched.columns:
            split_mm = pd.to_numeric(matched["__SPLIT__"], errors="coerce")
            split_mm.index = height_mm_final.index
            h_float = height_mm_final.astype(float)
            mask = split_mm.notna() & (split_mm > 0) & h_float.notna() & (h_float > 0)
            if mask.any():
                rounded = np.ceil(h_float[mask] / split_mm[mask]) * split_mm[mask]
                changed = (rounded != h_float[mask]).sum()
                log_message(f"Slat-rounding: {int(mask.sum())} shutter rows have חלוקה; {int(changed)} heights were rounded up.")
                for idx in h_float[mask].index:
                    log_message(f"  idx={idx}: height {h_float[idx]:.0f} → {rounded[idx]:.0f} (חלוקה={split_mm[idx]:.0f})")
                h_float.loc[mask] = rounded
                height_mm_final = h_float

        area_m2_out = (width_mm.astype(float)/1000.0 * (height_mm_final.astype(float)/1000.0)).round(5)

        project = project_from_stem(src_path.stem)

        # FIXED: Better color detection - check for non-empty strings, not just non-NA
        log_message(f"=== COLOR DETECTION DEBUG ===")
        log_message(f"matched dataframe has {len(matched)} rows")
        log_message(f"matched columns: {matched.columns.tolist()}")

        color_series = None
        # Try in order of priority: external color first, then internal, then aluminum shade, then generic shade
        for color_col in ["צבע חיצוני", "צבע פנימי", "גוון אלומיניום", "גוון"]:
            if color_col in matched.columns:
                temp_series = matched[color_col]
                # CRITICAL FIX: Check for non-empty strings, not just non-NA
                # Empty strings are not NA, so we need to check both
                has_data = (temp_series.notna() & (temp_series.astype(str).str.strip() != '')).any()
                log_message(f"Checking column '{color_col}': exists=True, has_data={has_data}, sample_values={temp_series.head(3).tolist()}")
                if has_data:
                    color_series = temp_series
                    log_message(f"✓ Using color from column: {color_col} (has {(temp_series.notna() & (temp_series.astype(str).str.strip() != '')).sum()} non-empty values)")
                    break
                else:
                    log_message(f"✗ Column '{color_col}' exists but is empty or all whitespace")
            else:
                log_message(f"Column '{color_col}' does not exist in matched dataframe")

        # NEW: If no color found in matched rows, use project_color as fallback
        if color_series is None or (color_series.astype(str).str.strip() == '').all():
            if project_color:
                log_message(f"⚡ No color in matched rows - applying project color '{project_color}' to all {len(matched)} rows")
                color_series = pd.Series([project_color] * len(matched), index=matched.index)
            else:
                log_message("⚠ WARNING: No color column found with data and no project color - גוון will be empty!")
                color_series = pd.Series([pd.NA] * len(matched), index=matched.index)
        else:
            log_message(f"Final color_series has {(color_series.notna() & (color_series.astype(str).str.strip() != '')).sum()} non-empty values")

        log_message(f"=== END COLOR DETECTION ===\n")

        # Get code
        code_series = None
        for c in ["קוד זיהוי","מס פרט","מספר פרט","מספר פריט","קוד פריט","קוד פנימי"]:
            if c in matched.columns:
                code_series = matched[c]
                log_message(f"Using code from column: {c}")
                break

        # FIXED: Build base output with PROPER COLUMN ORDER
        # Order: פרויקט, קוד זיהוי, שם-מקייט, תאור, ספק, גוון, רוחב, גובה, שטח, כמות
        base = pd.DataFrame({
            "פרויקט": project,
            "קוד זיהוי": (code_series.astype(str).str.replace("^V","", regex=True)
                           if code_series is not None else pd.NA),
            "שם-מקייט": pd.to_numeric(matched["__SKU__"], errors="coerce").round(0).astype("Int64"),
            "תאור": matched["__DESC__"],
            "ספק": matched.get("__VENDOR__"),
            "גוון": color_series,
            "רוחב": width_mm.round(0).astype("Int64"),
            "גובה": height_mm_final.round(0).astype("Int64"),
            "שטח ליחידה במ\"ר": area_m2_out,
            "_area_input_temp": area_m2_in,  # Temporary column for quantity calculation
        })

        # Calculate quantity
        # FIXED: Correct formula is total_area / area_per_unit, not the inverse!
        # For rows where input area exists, use it; otherwise default to 1
        with np.errstate(divide='ignore', invalid='ignore'):
            area_per_unit = (base["רוחב"].astype(float)/1000.0) * (base["גובה"].astype(float)/1000.0)
            has_input_area = base["_area_input_temp"].notna() & (base["_area_input_temp"] > 0)
            qty = pd.Series([pd.NA] * len(base), index=base.index)
            # Use input area / area_per_unit when available
            qty[has_input_area] = base.loc[has_input_area, "_area_input_temp"] / area_per_unit[has_input_area]
            # Default to 1 when no input area
            qty[~has_input_area] = 1
        base["כמות"] = pd.to_numeric(qty, errors="coerce").round(0).astype("Int64")

        # Remove temporary column
        base = base.drop(columns=["_area_input_temp"])

        # Validation
        valid_mask = base["גובה"].between(300,6000, inclusive="both") & base["רוחב"].between(300,6000, inclusive="both")
        bad = base[~valid_mask].copy()
        bad["סיבה"] = "גובה/רוחב מחוץ לטווח (300–6000 מ\"מ) גם אחרי תיקון יחידות"
        good = base[valid_mask].copy()

        log_message(f"Valid rows: {len(good)}, Invalid rows: {len(bad)}")

        # Code selection
        # FIXED v1.4h40: Pattern now accepts letters, dots, parentheses, and Hebrew characters
        # Examples: "12", "12L", "12.1L", "A1(A)", "B2(C)", "14.2C" - all valid
        pattern = re.compile(r'^[\d\.A-Za-z\u0590-\u05FF\(\)]{1,20}$')
        uniq_codes = sorted({ s for s in good["קוד זיהוי"].astype(str)
                              if s and s.lower()!='nan' and len(s)<=20 and pattern.match(s)})

        # NEW: Filter out codes for shutters that have "חלוקה" in the mapping file
        # These are already classified as slat types and shouldn't have bottom rail calculated
        codes_with_split = set()
        if "__SPLIT__" in mp.columns:
            # Get SKUs from matched rows
            matched_skus = matched["__SKU__"].dropna().unique()
            # Find which ones have חלוקה defined
            skus_with_split = mp[mp["__SPLIT__"].notna() & (mp["__SPLIT__"].astype(str).str.strip() != '')]["__SKU__"].tolist()
            if skus_with_split:
                log_message(f"Found {len(skus_with_split)} SKUs with 'חלוקה' defined: {skus_with_split}")
                # Don't ask about bottom rail for these shutters - they're already slat types
                log_message("Skipping bottom rail calculation for slat-type shutters (those with 'חלוקה')")
                # Filter: only show codes that DON'T have חלוקה
                codes_to_show = [code for code in uniq_codes
                                if code not in codes_with_split]
                log_message(f"Offering bottom rail selection for {len(codes_to_show)} codes (out of {len(uniq_codes)} total)")
                uniq_codes = codes_to_show

        chosen = select_codes_from_list(uniq_codes) if uniq_codes else []

        bottom_rows=[]; total_len_m=0.0
        if chosen:
            sel = good[good["קוד זיהוי"].isin(chosen)].copy()
            if not sel.empty:
                length_m = (sel["רוחב"].astype(float) * sel["כמות"].astype(float)) / 1000.0
                total_len_m = float(np.nansum(length_m.values))
                br = sel.copy()
                br["תאור"] = "שלב משוך תחתון"
                br["כמות"] = pd.Series(length_m.round(3)).astype(object)
                br["גובה"] = pd.NA  # Clear height for bottom rail rows
                br["שטח ליחידה במ\"ר"] = pd.NA  # Clear area for bottom rail rows (measured in meters, not m²)
                bottom_rows.append(br)
                log_message(f"Added bottom rail rows for codes: {chosen}, total length: {total_len_m}m")

        out = good.copy()
        if bottom_rows: out = pd.concat([out]+bottom_rows, ignore_index=True)
        if total_len_m>0:
            summary = {k: pd.NA for k in out.columns}
            summary["פרויקט"]=project
            summary["תאור"]="סיכום אורך כולל שלב משוך (מ')"
            summary["כמות"]=round(total_len_m,3)
            out = pd.concat([out, pd.DataFrame([summary])], ignore_index=True)

        # Write output to selected directory
        out_path = output_dir / f"{src_path.stem} - פלט.xlsx"
        with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
            out.to_excel(xw, index=False, sheet_name="רכש")
            if not bad.empty:
                bad.to_excel(xw, index=False, sheet_name="שורות בעייתיות")

            # NEW: Set RTL layout and swap width/height columns
            log_message("Setting RTL layout and swapping width/height columns")
            for sheet_name in xw.book.sheetnames:
                ws = xw.book[sheet_name]
                # Set RTL layout
                ws.sheet_view.rightToLeft = True

                # Swap width/height columns
                # Find column indices for רוחב and גובה
                if ws.max_row > 0:
                    header_row = [cell.value for cell in ws[1]]
                    try:
                        width_col_idx = header_row.index("רוחב") + 1  # openpyxl is 1-indexed
                        height_col_idx = header_row.index("גובה") + 1

                        log_message(f"Swapping columns: רוחב (col {width_col_idx}) ↔ גובה (col {height_col_idx})")

                        # Swap column contents (including header)
                        for row_idx in range(1, ws.max_row + 1):
                            width_cell = ws.cell(row=row_idx, column=width_col_idx)
                            height_cell = ws.cell(row=row_idx, column=height_col_idx)

                            # Swap values
                            width_cell.value, height_cell.value = height_cell.value, width_cell.value

                        log_message(f"Successfully swapped width/height for sheet '{sheet_name}'")
                    except (ValueError, IndexError) as e:
                        log_message(f"Could not find רוחב/גובה columns in sheet '{sheet_name}': {e}")

                # Auto-fit column widths so Hebrew descriptions don't truncate.
                if ws.max_row > 0 and ws.max_column > 0:
                    for col_idx in range(1, ws.max_column + 1):
                        col_letter = get_column_letter(col_idx)
                        max_len = 0
                        for row_idx in range(1, ws.max_row + 1):
                            v = ws.cell(row=row_idx, column=col_idx).value
                            if v is not None:
                                vl = len(str(v))
                                if vl > max_len: max_len = vl
                        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 50)

        log_message(f"Output written successfully: {out_path}")
        log_message(f"Output shape: {out.shape}")

        try:
            root=tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
            msg = f"נוצר קובץ (v{APP_VERSION}):\n{out_path}\nוכן נשמר עותק UTF-8: {utf8_copy.name}"
            if not bad.empty:
                msg += f"\n\nאזהרה: הועברו {len(bad)} שורות לגיליון 'שורות בעייתיות'."
            messagebox.showinfo("הצלחה", msg); root.destroy()
        except Exception:
            print("Done:", out_path)

    except Exception as e:
        log_message(f"ERROR: {e}")
        import traceback
        log_message(traceback.format_exc())
        raise

if __name__ == "__main__":
    try:
        main()
    except SystemExit:
        pass
    except Exception as e:
        try:
            root=tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
            messagebox.showerror("שגיאה", f"{e}"); root.destroy()
        except Exception:
            print("שגיאה:", e)
