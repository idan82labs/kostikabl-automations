#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
aluminum_cleaner_v2.py
Fixes vs previous version:
- get_column_letter() replaces chr(64+idx) — works beyond column Z
- SKU match threshold raised from 0.05 → 0.3
- Removed "תאור" from POSSIBLE_COLOR_HEADERS (was causing wrong column detection)
- on_bad_lines default changed from "error" → "warn"
- Column-shift hack replaced with index_col=False in read_csv
- Encoding/sep loop prioritizes cp1255+semicolon first, then fallback
- Removed dead function has_letter_and_digit
"""

import argparse, sys, re, os
from typing import List, Optional, Set
import pandas as pd
from openpyxl.utils import get_column_letter

APP_VERSION = "2.3"

SHUTTER_SKU_PATTERN = re.compile(r"^(11|13)\d{6}$")
MAPPING_FILENAMES = ["kostika_mapping.xlsx", "מקטים תריסים כולל.xlsx"]


def _load_shutter_skus(input_path: str) -> Set[str]:
    """Return the set of SKUs (digits-only) that the master mapping classifies
    as shutter components. Searches PyInstaller's --onefile temp dir, next to
    the .exe (--onedir), and the input file's folder. Missing mapping is silent
    — the regex fallback still catches 11/13-prefixed SKUs."""
    exe_dir = os.path.dirname(os.path.abspath(sys.executable if getattr(sys, "frozen", False) else __file__))
    input_dir = os.path.dirname(input_path) or "."
    search_dirs = []
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        search_dirs.extend([meipass, os.path.join(meipass, "data")])
    search_dirs.extend([exe_dir, os.path.join(exe_dir, "data"), input_dir])
    candidates = [os.path.join(d, name) for d in search_dirs for name in MAPPING_FILENAMES]
    for p in candidates:
        if os.path.exists(p):
            try:
                mp = pd.read_excel(p, dtype=str)
                return {re.sub(r"\D", "", str(v)) for v in mp.iloc[:, 0].dropna() if str(v).strip()}
            except Exception:
                return set()
    return set()

try:
    import tkinter as tk
    from tkinter import filedialog
    TK_AVAILABLE = True
except Exception:
    TK_AVAILABLE = False

POSSIBLE_SKU_HEADERS = [
    "שם-מקייט", "שם-מקט", "שם מק\"ט", "שם-מק\"ט", "שםמקט",
    "sku", "מק\"ט", "מקט", "מק׳ט", "מק״ט", "מק-ט", "kat", "מ.ק.ט", "מק ט"
]

# Removed "תאור" — it's a description field, not a color field
POSSIBLE_COLOR_HEADERS = [
    "גוון", "צבע", "shade", "color", "ral",
    "קוד גוון", "קוד צבע", "צבע פנימי", "צבע חיצוני"
]

SKU_PATTERN = re.compile(r"^\d{4,}$")


def normalize_header(h: str) -> str:
    h2 = str(h).strip().lower()
    h2 = h2.replace("״", "\"").replace("׳", "'").replace("\u2019", "'").replace("`", "'")
    h2 = re.sub(r"\s+", " ", h2.replace("-", " ").replace("_", " ")).strip()
    return h2


def autodetect_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    norm_map = {col: normalize_header(col) for col in df.columns}
    cand_norms = [normalize_header(c) for c in candidates]
    for col, norm in norm_map.items():
        if norm in cand_norms:
            return col
    for col, norm in norm_map.items():
        for cn in cand_norms:
            if cn and cn in norm:
                return col
    return None


def pick_sku_like_column(df: pd.DataFrame) -> Optional[str]:
    best_col, best_ratio = None, 0.0
    for col in df.columns:
        s = df[col].astype(str)  # NaN → 'nan' → has no digits, safely scores 0
        ratio = s.map(lambda x: bool(re.fullmatch(r"\d{4,}", re.sub(r"\D", "", x)))).mean()
        if ratio > best_ratio:
            best_ratio, best_col = ratio, col
    return best_col if best_ratio >= 0.3 else None  # Raised from 0.05 → 0.3


def _sku_match_ratio(series: pd.Series) -> float:
    if series is None or len(series) == 0:
        return 0.0
    return series.map(
        lambda x: bool(re.fullmatch(r"\d{4,}", re.sub(r"\D", "", str(x) if x is not None else "")))
    ).mean()


def _smart_choose_sku(df: pd.DataFrame, preferred: Optional[str]) -> str:
    if preferred and preferred in df.columns and _sku_match_ratio(df[preferred]) >= 0.05:
        return preferred
    picked = autodetect_column(df, POSSIBLE_SKU_HEADERS)
    if picked and _sku_match_ratio(df[picked]) >= 0.05:
        return picked
    cand = pick_sku_like_column(df)
    if cand:
        return cand
    return preferred if (preferred and preferred in df.columns) else (picked if picked else df.columns[0])


def try_read_csv_or_excel(path: str, encoding_hint: Optional[str], sep_hint: Optional[str], on_bad_lines: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext in {".xlsx", ".xls"}:
        return pd.read_excel(path, dtype=str)

    # Prioritized attempts: most common case first
    priority = [("cp1255", ";"), ("utf-8-sig", ";"), ("utf-8", ";")]
    if encoding_hint or sep_hint:
        priority.insert(0, (encoding_hint or "cp1255", sep_hint or ";"))

    fallback = [
        ("cp1255", ","), ("utf-8-sig", ","), ("utf-8", ","),
        ("iso-8859-8", ";"), ("iso-8859-8", ","),
        ("cp1255", "\t"), ("utf-8", "\t"),
    ]

    last_err = None
    for enc, sep in priority + fallback:
        try:
            return pd.read_csv(
                path, encoding=enc, sep=sep, engine="python",
                on_bad_lines=on_bad_lines, index_col=False, dtype=str  # index_col=False fixes column shift
            )
        except Exception as e:
            last_err = e

    # Last resort
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return pd.read_csv(f, sep=None, engine="python", on_bad_lines=on_bad_lines, index_col=False, dtype=str)


def rule_sku_is_numeric(sku_val) -> bool:
    return bool(SKU_PATTERN.match(re.sub(r"\D", "", str(sku_val) if sku_val is not None else "")))


def rule_color_nonempty(color_val) -> bool:
    s = "" if (color_val is None or (isinstance(color_val, float) and pd.isna(color_val))) else str(color_val)
    return len(s.strip()) > 0


def reconcile_with_xls(df: pd.DataFrame, xls_path: str) -> pd.DataFrame:
    """Replace `כמות` and `משקל` in df using the XLS qty source.

    For each SKU, scale every CSV row by xls_qty / sum(csv_qty for that SKU).
    This handles single-row and multi-row-per-SKU cases identically and keeps
    the row split intact while making per-SKU totals match the XLS.

    SKUs only in CSV are left as-is. SKUs only in XLS are dropped (cannot
    derive kg-per-unit without a CSV reference row)."""
    xls = pd.read_excel(xls_path, dtype=str)

    xls_sku_col = autodetect_column(xls, POSSIBLE_SKU_HEADERS) or xls.columns[0]
    xls_qty_col = next(
        (c for c in xls.columns if normalize_header(c) == "כמות" or "כמות" in str(c)),
        None,
    )
    if not xls_qty_col:
        sys.stderr.write("WARN: 'כמות' column not found in XLS — skipping reconciliation\n")
        return df

    xls_map = {}
    for _, row in xls.iterrows():
        sku = re.sub(r"\D", "", str(row[xls_sku_col] or ""))
        if not sku:
            continue
        try:
            q = float(row[xls_qty_col])
        except (TypeError, ValueError):
            continue
        if q <= 0:
            continue
        xls_map[sku] = q

    csv_qty_col = next((c for c in df.columns if normalize_header(c) == "כמות"), None)
    csv_weight_col = next((c for c in df.columns if normalize_header(c) == "משקל"), None)
    if not csv_qty_col:
        sys.stderr.write("WARN: CSV missing 'כמות' column — skipping reconciliation\n")
        return df

    csv_sku_col = _smart_choose_sku(df, None)

    df = df.copy()
    df[csv_qty_col] = df[csv_qty_col].astype(object)
    if csv_weight_col is not None:
        df[csv_weight_col] = df[csv_weight_col].astype(object)
    sku_norm = df[csv_sku_col].astype(str).map(lambda x: re.sub(r"\D", "", x))
    csv_qty_num = pd.to_numeric(df[csv_qty_col], errors="coerce").fillna(0.0)
    csv_weight_num = pd.to_numeric(df[csv_weight_col], errors="coerce") if csv_weight_col else None
    sums = csv_qty_num.groupby(sku_norm).sum().to_dict()

    matched = 0
    csv_only = []
    for idx in df.index:
        sku = sku_norm.loc[idx]
        if not sku:
            continue
        if sku not in xls_map:
            csv_only.append(sku)
            continue
        csv_sum = sums.get(sku, 0.0)
        xls_qty = xls_map[sku]
        if csv_sum > 0:
            ratio = xls_qty / csv_sum
            df.at[idx, csv_qty_col] = csv_qty_num.loc[idx] * ratio
            if csv_weight_col is not None:
                df.at[idx, csv_weight_col] = csv_weight_num.loc[idx] * ratio
        else:
            n_rows = int((sku_norm == sku).sum())
            df.at[idx, csv_qty_col] = xls_qty / max(1, n_rows)
            if csv_weight_col is not None:
                df.at[idx, csv_weight_col] = None
        matched += 1

    csv_skus = set(sku_norm)
    xls_only = [s for s in xls_map if s not in csv_skus]
    if xls_only:
        sys.stderr.write(f"XLS-only SKUs dropped (no CSV reference for kg-per-unit): {xls_only}\n")
    csv_only_unique = sorted(set(csv_only))
    if csv_only_unique:
        head = csv_only_unique[:10]
        suffix = f" (+{len(csv_only_unique) - 10} more)" if len(csv_only_unique) > 10 else ""
        sys.stderr.write(f"CSV-only SKUs kept as-is: {head}{suffix}\n")
    sys.stderr.write(f"Reconciliation: scaled {matched} rows against XLS quantities\n")
    return df


def pick_files_interactively():
    if TK_AVAILABLE:
        root = tk.Tk(); root.withdraw()
        inp = filedialog.askopenfilename(
            title="בחר/י קובץ קלט CSV (מקור Priority)",
            filetypes=[("CSV / Excel", "*.csv *.xlsx *.xls")]
        )
        if not inp: sys.exit(0)
        xls_qty = filedialog.askopenfilename(
            title="בחר/י קובץ XLS עם כמויות מעודכנות (Cancel = מצב CSV בלבד)",
            filetypes=[("XLS / Excel", "*.xls *.xlsx")]
        ) or None
        base, _ = os.path.splitext(os.path.basename(inp))
        initialdir = os.path.dirname(inp) or os.getcwd()
        out = filedialog.asksaveasfilename(
            title="בחר/י היכן לשמור פלט (Excel)",
            defaultextension=".xlsx",
            initialfile=f"{base}.cleaned.xlsx",
            initialdir=initialdir,
            filetypes=[("Excel", "*.xlsx")]
        )
        if not out: sys.exit(0)
        return inp, out, xls_qty

    inp = input("Enter input file path: ").strip('"')
    base = os.path.splitext(os.path.basename(inp))[0]
    out_default = os.path.join(os.path.dirname(inp) or ".", f"{base}.cleaned.xlsx")
    out = input(f"Enter output Excel path [default: {out_default}]: ").strip('"') or out_default
    xls_qty = input("Optional XLS qty-update path (blank = none): ").strip('"') or None
    return inp, out, xls_qty


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input_path", nargs="?")
    ap.add_argument("output_csv", nargs="?")
    ap.add_argument("--sku-col", dest="sku_col")
    ap.add_argument("--color-col", dest="color_col")
    ap.add_argument("--columns", dest="columns")
    ap.add_argument("--print-stats", action="store_true")
    ap.add_argument("--sep")
    ap.add_argument("--encoding")
    ap.add_argument("--on-bad-lines", default="warn", choices=["error", "skip", "warn"])  # default changed
    ap.add_argument("--qty-update", dest="qty_update", default=None,
                    help="Optional XLS file with updated quantities; weight is recomputed proportionally")
    args = ap.parse_args()

    if not args.input_path and not args.output_csv:
        input_path, output_csv, qty_update = pick_files_interactively()
    elif args.input_path and not args.output_csv:
        stem, _ = os.path.splitext(os.path.basename(args.input_path))
        output_csv = os.path.join(os.path.dirname(args.input_path) or ".", f"{stem}.cleaned.xlsx")
        input_path = args.input_path
        qty_update = args.qty_update
    else:
        input_path, output_csv = args.input_path, args.output_csv
        qty_update = args.qty_update

    df = try_read_csv_or_excel(input_path, args.encoding, args.sep, args.on_bad_lines)

    # Remove unnamed/empty columns (from trailing delimiters)
    df = df.drop(columns=[c for c in df.columns if str(c).startswith("Unnamed") or str(c).strip() == ""])

    if qty_update:
        df = reconcile_with_xls(df, qty_update)

    # Sanity-check: catch the common operator mistake of passing the qty-update XLS
    # as the primary input. The rich Priority CSV always has 'פרויקט' AND 'מידה';
    # the 'מחיר לסדרה' XLS has neither.
    required_cols = ["פרויקט", "מידה"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        sys.stderr.write(
            f"ERROR: Input is missing required columns: {missing}\n"
            f"  Input columns ({len(df.columns)}): {df.columns.tolist()[:15]}{'...' if len(df.columns) > 15 else ''}\n"
            f"  HINT: The main input must be the rich Priority CSV. The 'מחיר לסדרה' XLS\n"
            f"  belongs in --qty-update, not as the primary input.\n"
        )
        sys.exit(2)

    sku_col = _smart_choose_sku(df, args.sku_col)

    color_col = args.color_col or autodetect_column(df, POSSIBLE_COLOR_HEADERS)
    if not color_col:
        for alt in ["צבע פנימי", "צבע חיצוני", "גוון", "צבע"]:
            if alt in df.columns:
                color_col = alt
                break
    if not color_col:
        cols = df.columns.tolist()
        sys.stderr.write(
            f"ERROR: Color/Shade column not found.\n"
            f"  Input has {len(cols)} columns: {cols[:15]}{'...' if len(cols) > 15 else ''}\n"
            f"  HINT: The main input must be the rich Priority CSV (with 'צבע פנימי' / 'גוון').\n"
            f"  If you meant the 'מחיר לסדרה' XLS quantity-update report, pass it via --qty-update,\n"
            f"  not as the primary input.\n"
        )
        sys.exit(2)

    rule1 = df[sku_col].apply(rule_sku_is_numeric)
    rule2 = df[color_col].apply(rule_color_nonempty)
    rule3 = pd.Series(True, index=df.index)
    if "מידה" in df.columns:
        rule3 = df["מידה"].notna() & (df["מידה"].astype(str).str.strip() != "")

    # Shutter-family check applied as a filter rule so --print-stats reflects the drop.
    shutter_skus = _load_shutter_skus(input_path)
    def _is_shutter(sku_val) -> bool:
        digits = re.sub(r"\D", "", str(sku_val) if sku_val is not None else "")
        return digits in shutter_skus or bool(SHUTTER_SKU_PATTERN.match(digits))
    rule_not_shutter = ~df[sku_col].apply(_is_shutter)

    # Log specifically which shutter rows would have otherwise passed the other rules.
    shutter_drops_mask = rule1 & rule2 & rule3 & ~rule_not_shutter
    n_shutter_dropped = int(shutter_drops_mask.sum())
    if n_shutter_dropped:
        dropped = df.loc[shutter_drops_mask, sku_col].astype(str).tolist()
        sys.stderr.write(f"Filtered {n_shutter_dropped} shutter rows from material output: {dropped}\n")

    kept = df[rule1 & rule2 & rule3 & rule_not_shutter].copy()

    desired_columns = [
        ("פרויקט",            ["פרויקט"]),
        ("שם-מקייט",          ["שם-מקייט"]),
        ("תאור",              ["תאור"]),
        ("גוון",              ["גוון", "צבע פנימי", "צבע חיצוני", "צבע"]),
        ("מידה",              ["מידה"]),
        ("כמות",              ["כמות"]),
        ("יחידת מידה",        ["יחידת מידה"]),
        ("משקל",              ["משקל"]),
        ("יחידת מידה מישקל",  ["יחידת מידה מישקל"]),
        ("ספק",               ["ספק"]),
    ]

    final_columns = []
    rename_dict = {}

    for output_name, possible_inputs in desired_columns:
        for input_name in possible_inputs:
            if input_name in kept.columns:
                final_columns.append(input_name)
                if input_name != output_name:
                    rename_dict[input_name] = output_name
                break
            input_norm = normalize_header(input_name)
            for actual_col in kept.columns:
                if normalize_header(actual_col) == input_norm:
                    final_columns.append(actual_col)
                    if actual_col != output_name:
                        rename_dict[actual_col] = output_name
                    break
            else:
                continue
            break

    kept = kept[final_columns].rename(columns=rename_dict)

    # Pad SKU to at least 5 digits
    if "שם-מקייט" in kept.columns:
        def pad_sku(val):
            digits = re.sub(r"\D", "", str(val))
            return digits.zfill(5) if len(digits) <= 4 else digits
        kept["שם-מקייט"] = kept["שם-מקייט"].apply(pad_sku)

    # Normalize output path
    if not output_csv.endswith(".xlsx"):
        output_xlsx = (output_csv[:-4] if output_csv.endswith(".csv") else output_csv) + ".xlsx"
    else:
        output_xlsx = output_csv

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        kept.to_excel(writer, index=False, sheet_name="Sheet1")
        ws = writer.sheets["Sheet1"]

        ws.sheet_view.rightToLeft = True

        for idx, col_name in enumerate(kept.columns, start=1):
            col_letter = get_column_letter(idx)
            content_max = max((len(str(v)) for v in kept[col_name].tolist() if v is not None), default=0)
            width = min(max(max(content_max, len(str(col_name))) + 2, 8), 50)
            ws.column_dimensions[col_letter].width = width
            if col_name == "שם-מקייט":
                for row in range(2, len(kept) + 2):
                    ws[f"{col_letter}{row}"].number_format = "@"

    if args.print_stats:
        removed = len(df) - len(kept)
        print(f"Rows in: {len(df)} | kept: {len(kept)} | removed: {removed} (incl. {n_shutter_dropped} shutter rows)")
        print(f"SKU column: {sku_col} | Color column: {color_col}")
        if qty_update:
            print(f"Qty-update XLS applied: {qty_update}")
        print(f"Saved: {output_xlsx}")
    else:
        print(f"נשמר: {output_xlsx}")


if __name__ == "__main__":
    main()
